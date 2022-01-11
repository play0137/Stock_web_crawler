""" 改變連續n個月被篩選出來的股票名稱的樣式 """
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import global_vars
import pdb

# file_path = global_vars.DIR_PATH + "stock_filter_history.xlsx"
file_path = global_vars.DIR_PATH + "月營收創新高_backup.xlsx"

dfs = pd.read_excel(file_path, sheet_name=None) # 設sheet_name為None會讀進excel的所有活頁簿
stocks_list = list() # 記錄所有創新高股票的名稱
for sheet_name in dfs:
    df = dfs[sheet_name]
    stocks_list.append(df["名稱"].tolist())

workbook_size = len(dfs)
while True:
    past_months, continuous_months = input("請輸入過去幾個月裡連續幾個月創新高:\n").split(",")
    past_months = int(past_months)
    continuous_months = int(continuous_months)
    if past_months > workbook_size or continuous_months > past_months or past_months < 0 or continuous_months < 0:
        print("請重新輸入")
        continue
    break

# 計算迴圈次數
loop_times = past_months - continuous_months + 1
start_month = (datetime.now().month-1) - past_months + 1
start_index = workbook_size - past_months
intersection_list = list()
for i in range(loop_times):
    intersection_list.append(list(set.intersection(*[set(s) for s in stocks_list[start_index:start_index+continuous_months]]))) # set.intersection:list裡的每個元素是set{}
    print(f"從{start_month}月到{start_month+continuous_months-1}月:\n{intersection_list[i]}")
    start_month += 1
    start_index += 1
# print(intersection_list)

# fill the background color if the cell value in the intersections
# pf = PatternFill(fill_type="solid", start_color="FFB5B5") # 前景:淡紅色(目前先不改前景顏色)
# pf_reset = PatternFill(fill_type="solid", start_color="FFFFFF") # 前景重置顏色:白色
ft = Font(color="FF0000") # 字體:紅色
ft_reset = Font(color="000000") # 字體重置顏色:黑色

sheet_start_index = workbook_size - past_months
wb = load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames
for i, sheet_name in enumerate(sheet_names):
    ws = wb[sheet_name] # ws:worksheet
    
    # add filter to excel
    ws.auto_filter.ref = "B1:B500"
    ws.auto_filter.add_sort_condition("B1:B500")
    
    if i >= sheet_start_index: # 只有在需要時才計算交集
        inter_start_index = i-sheet_start_index+1-continuous_months #計算intersection_list的起點
        if inter_start_index < 0: # start_index不能為負數
            inter_start_index = 0
        intersection = intersection_list[inter_start_index:i+1]
        intersection = set.union(*[set(s) for s in intersection]) # 同一個月份可能出現在不同的連續月份裡
        print(f"{sheet_name}:{intersection}")
    """
        例如:在過去5個月(6,7,8,9,10月)連續3個月出現
        假設
        6,7,8月重複的資料:[1, 2, 3, 5]
        7,8,9月重複的資料:[1, 2, 5, 11]
        8,9,10月重複的資料:[2, 5]
        
        各月份要標記的資料如下列:
        6月:{1, 2, 3, 5}
        7月:{1, 2, 3, 5, 11}
        8月:{1, 2, 3, 5, 11}
        9月:{1, 2, 5, 11}
        10月:{2, 5}
        
        8月會同時出現在(6,7,8月), (7,8,9月), (8,9,10月)
        不能只憑靠6,7,8月重複資料:[1, 2, 3, 5]來決定, 還少了11
        所以要將(6,7,8月), (7,8,9月), (8,9,10月)的資料聯集(union)起來
    """
    # for row in ws.iter_rows(min_row=ws.min_row+1, max_row=ws.max_row):
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row):
        for cell in row:
            if cell.col_idx == 1: # 只reset"名稱"那一欄
                # cell.fill = pf_reset
                cell.font = ft_reset
            if i >= sheet_start_index and cell.value in intersection:
                # cell.fill = pf
                cell.font = ft
wb.save(file_path)