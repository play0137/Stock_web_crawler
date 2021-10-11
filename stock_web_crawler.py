"""
stock web crawler
crawl the stock information from the Goodinfo! website
1.
GoodInfo -> 股票篩選 -> 智慧選股 -> 月營收 -> 月營收創排名紀錄 -> 單月營收創歷史新高
2.
GoodInfo -> 股票篩選 -> 智慧選股 -> 獲利狀況 -> 營業毛利創新高/低 -> 單季營業毛利創歷季新高
                                               營業利益創新高/低 -> 單季營業利益創歷季新高
                                               稅後淨利創新高/低 -> 單季稅後淨利創歷季新高
3.
GoodInfo -> 每月營收 -> 輸入某個 股票代號/名稱 -> 抓取過去各年月的歷史資料(營業收入部分, 單月及累計下的資料)
4.
GoodInfo -> 每月營收 -> 輸入某個 股票代號/名稱 -> 經營績效 -> 合併報表-單季 -> 抓下面的table -> 放在個股月營收excel檔下方
https://goodinfo.tw/StockInfo/StockBzPerformance.asp?STOCK_ID=2330&YEAR_PERIOD=9999&RPT_CAT=M_QUAR
"""

# standard libraries
import os
import sys
import pdb
import time
import random
import requests
from datetime import datetime

# third-party libraries
import pandas as pd
import numpy as np
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from msedge.selenium_tools import Edge, EdgeOptions
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import Workbook

import global_vars

def main():
    global_vars.initialize_proxy()
    
    # webdriver
    options = EdgeOptions()
    options.use_chromium = True
    options.add_argument (f"--user-agent={UserAgent().random}")
    options.add_argument("--incognito")           # incognito mode  
    options.add_argument("headless")              # executing selenium without running the browser
    options.add_argument("disable-notifications") # disable notifications
    driver = Edge(global_vars.DIR_PATH + "edgedriver_win64/msedgedriver.exe", options=options)
    driver.implicitly_wait(random.randint(1,4))
    
    """ 月營收創新高 """
    stock_highest_salemon_file = global_vars.DIR_PATH + "月營收創新高.xlsx"
    url = "https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E6%99%BA%E6%85%A7%E9%81%B8%E8%82%A1&INDUSTRY_CAT=%E5%96%AE%E6%9C%88%E7%87%9F%E6%94%B6%E5%89%B5%E6%AD%B7%E5%8F%B2%E6%96%B0%E9%AB%98%40%40%E6%9C%88%E7%87%9F%E6%94%B6%E5%89%B5%E6%8E%92%E5%90%8D%E7%B4%80%E9%8C%84%40%40%E5%96%AE%E6%9C%88%E7%87%9F%E6%94%B6%E5%89%B5%E6%AD%B7%E5%8F%B2%E6%96%B0%E9%AB%98"
    table_ID = "#divStockList"
    df = stock_crawler(url, None, table_ID)
    delete_header(df, list(df.columns))
    convert_dtype(df) # convert numeric values in string to float
    
    sheet_name = f"{datetime.now().month-1}月"         # write to different months
    if not os.path.isfile(stock_highest_salemon_file): # create an excel file if not exist
        wb = Workbook()
        wb.worksheets[0].title = sheet_name
        wb.save(stock_highest_salemon_file)
    writer = pd.ExcelWriter(stock_highest_salemon_file, mode="a", engine="openpyxl", if_sheet_exists='replace')
    df = df[df["營收月份"] == f"{datetime.now().year%100}M{datetime.now().month-1:02d}"] # only save month-1 data
    if df:
        df.to_excel(writer, index=False, encoding="UTF-8", sheet_name=sheet_name, freeze_panes=(1,2))
        excel_formatting(writer, df, sheet_name)
        writer.save()
    
    """ drop down menu """
    stock_crawler_file = global_vars.DIR_PATH + "stock_crawler.xlsx"
    sheet_name_list = ["負債比", "全體董監持股_全體董監質押比", "本國法人持股", "單季營業毛利創歷季新高", "單季營業利益創歷季新高", "單季稅後淨利創歷季新高", "殖利率", "均線"]
    url_list = list()
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E8%B2%A0%E5%82%B5%E7%B8%BD%E9%A1%8D%E4%BD%94%E7%B8%BD%E8%B3%87%E7%94%A2%E6%AF%94%E6%9C%80%E9%AB%98%40%40%E8%B2%A0%E5%82%B5%E7%B8%BD%E9%A1%8D%40%40%E8%B2%A0%E5%82%B5%E7%B8%BD%E9%A1%8D%E4%BD%94%E7%B8%BD%E8%B3%87%E7%94%A2%E6%AF%94%E6%9C%80%E9%AB%98")
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%85%A8%E9%AB%94%E8%91%A3%E7%9B%A3%E6%8C%81%E8%82%A1%E6%AF%94%E4%BE%8B%28%25%29%40%40%E5%85%A8%E9%AB%94%E8%91%A3%E7%9B%A3%40%40%E6%8C%81%E8%82%A1%E6%AF%94%E4%BE%8B%28%25%29")
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E6%9C%AC%E5%9C%8B%E6%B3%95%E4%BA%BA%E6%8C%81%E8%82%A1%E6%AF%94%E4%BE%8B%28%25%29%40%40%E6%9C%AC%E5%9C%8B%E6%B3%95%E4%BA%BA%40%40%E6%8C%81%E8%82%A1%E6%AF%94%E4%BE%8B%28%25%29")
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E6%99%BA%E6%85%A7%E9%81%B8%E8%82%A1&INDUSTRY_CAT=%E5%96%AE%E5%AD%A3%E7%87%9F%E6%A5%AD%E6%AF%9B%E5%88%A9%E5%89%B5%E6%AD%B7%E5%AD%A3%E6%96%B0%E9%AB%98%40%40%E7%87%9F%E6%A5%AD%E6%AF%9B%E5%88%A9%E5%89%B5%E6%96%B0%E9%AB%98%2F%E4%BD%8E%40%40%E5%96%AE%E5%AD%A3%E7%87%9F%E6%A5%AD%E6%AF%9B%E5%88%A9%E5%89%B5%E6%AD%B7%E5%AD%A3%E6%96%B0%E9%AB%98")
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E6%99%BA%E6%85%A7%E9%81%B8%E8%82%A1&INDUSTRY_CAT=%E5%96%AE%E5%AD%A3%E7%87%9F%E6%A5%AD%E5%88%A9%E7%9B%8A%E5%89%B5%E6%AD%B7%E5%AD%A3%E6%96%B0%E9%AB%98%40%40%E7%87%9F%E6%A5%AD%E5%88%A9%E7%9B%8A%E5%89%B5%E6%96%B0%E9%AB%98%2F%E4%BD%8E%40%40%E5%96%AE%E5%AD%A3%E7%87%9F%E6%A5%AD%E5%88%A9%E7%9B%8A%E5%89%B5%E6%AD%B7%E5%AD%A3%E6%96%B0%E9%AB%98")
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E6%99%BA%E6%85%A7%E9%81%B8%E8%82%A1&INDUSTRY_CAT=%E5%96%AE%E5%AD%A3%E7%A8%85%E5%BE%8C%E6%B7%A8%E5%88%A9%E5%89%B5%E6%AD%B7%E5%AD%A3%E6%96%B0%E9%AB%98%40%40%E7%A8%85%E5%BE%8C%E6%B7%A8%E5%88%A9%E5%89%B5%E6%96%B0%E9%AB%98%2F%E4%BD%8E%40%40%E5%96%AE%E5%AD%A3%E7%A8%85%E5%BE%8C%E6%B7%A8%E5%88%A9%E5%89%B5%E6%AD%B7%E5%AD%A3%E6%96%B0%E9%AB%98"    )
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E7%8F%BE%E9%87%91%E6%AE%96%E5%88%A9%E7%8E%87+%28%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6%29%40%40%E7%8F%BE%E9%87%91%E6%AE%96%E5%88%A9%E7%8E%87%40%40%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6")
    url_list.append("https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E6%99%BA%E6%85%A7%E9%81%B8%E8%82%A1&INDUSTRY_CAT=%E6%9C%88K%E7%B7%9A%E7%AA%81%E7%A0%B4%E5%AD%A3%E7%B7%9A%40%40%E6%9C%88K%E7%B7%9A%E5%90%91%E4%B8%8A%E7%AA%81%E7%A0%B4%E5%9D%87%E5%83%B9%E7%B7%9A%40%40%E5%AD%A3%E7%B7%9A")
    
    writer = pd.ExcelWriter(stock_crawler_file, mode="w", engine="xlsxwriter")
    dropdown_ID = "selRANK"    # 下拉式選單ID
    table_ID = "#divStockList" # 表格ID
    for i in range(len(sheet_name_list)):
        driver.get(url_list[i])
        try:
            driver.find_element_by_id(dropdown_ID) # to test whether the dropdown_ID is in the website or not
            df = stock_crawler_dropdown(driver, dropdown_ID, table_ID)
        except NoSuchElementException:
            df = stock_crawler(None, driver.page_source, table_ID)
        delete_header(df, list(df.columns))
        # df.drop_diplicates(inplace=True)
        df.to_excel(writer, index=False, encoding="UTF-8", sheet_name=sheet_name_list[i], freeze_panes=(1,2))
        excel_formatting(writer, df, sheet_name_list[i])
    writer.save()
    
    
# delete duplicate header in data
def delete_header(df, header):
    first_header = header[0]
    # skip all the indexes which are the same as the first header
    skip_rows = df[df[df.columns[0]]==first_header].index
    df.drop(skip_rows, inplace=True)

# convert numeric values in string dtype of dataframe to float
def convert_dtype(df):
    for name in df.columns:
        if not np.isnan(pd.to_numeric(df[name], errors='coerce')).any(): # coerce:invalid parsing will be set as NaN
            df[name] = df[name].astype(float)

def stock_crawler(url, page_source, table_ID, table_number=0):
    while True:
        if page_source:
            soup = BeautifulSoup(page_source, "lxml")
        else:
            # fake user agent
            fake_ua = UserAgent()
            header = {'User-Agent': fake_ua.random}
            # header = {"User-Agent": "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.93 Safari/537.36"}
            
            # request
            if not global_vars.proxy_list:
                response = requests.post(url, headers=header)
            else:
                response = requests.post(url, headers=header, proxies=global_vars.proxy)
            if response.status_code != requests.codes.ok: # status_code:200
                sys.stderr.write("Request failed!\n")
                sys.stderr.write("Status code:" + str(response.status_code) + '\n')
                sys.stderr.write("Site:" + url + '\n')
            soup = BeautifulSoup(response.content, "lxml")
        if "異常" in soup.text or "請勿透過網站內容下載" in soup.text:
            sys.stderr.write("異常下載\n")
            global_vars.update_proxy()
            time.sleep(random.randint(30,60)) # sleep n seconds
            continue
        print("Request successful!")
        break
    
    soup.encoding = "UTF-8"
    try:
        div = soup.select_one(table_ID)
        if table_number == -1:
            df = pd.read_html(str(div))
        else:
            df = pd.read_html(str(div))[table_number]
    except:
        pdb.set_trace()
        sys.exit("empty div")
        
    return df

def stock_crawler_dropdown(driver, dropdown_ID, table_ID):
    element = driver.find_element_by_id(dropdown_ID)
    options_num = len(element.text.split('\n'))
    for i in range(options_num):
        try:
            select = Select(element).options[i]
            select.click()
        except:
            # refetch if the element is no longer attached to the DOM
            element = driver.find_element_by_id(dropdown_ID) # 1~300, 301~600, 601~900, 901~1200, 1201~1500, 1500~1734
            select = Select(element).options[i]
            select.click()
        time.sleep(random.randint(2,5))
        df = stock_crawler(None, driver.page_source, table_ID)
        if i == 0:
            df_concat = df
        else:
            df_concat = pd.concat([df_concat, df], axis=0)
            
    return df_concat

# excel format setting
def excel_formatting(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]

    # align to center
    if writer.engine == "openpyxl":
        for row_cells in worksheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    elif writer.engine == "xlsxwriter":
        cell_format = writer.book.add_format()
        cell_format.set_align('center')
        cell_format.set_align('vcenter')
        for col, col_name in enumerate(df.columns):
            worksheet.set_column(col, col, None, cell_format)
            
    set_column_width(worksheet, df, writer.engine)
        
# set column width to the max length of column cells
def set_column_width(worksheet, df, engine):
    for col, col_name in enumerate(df.columns):
        df_list = list(df[col_name].astype(str))
        max_col_len = -1
        for cell in df_list:
            if len_byte(cell) > max_col_len:
                max_col_len = len_byte(cell)
        max_col_len = int(max(max_col_len, len_byte(col_name)))
        if engine == "openpyxl":
            worksheet.column_dimensions[get_column_letter(col+1)].width = max_col_len # arg in column_dimensions is letter('A->1, 'B'->2, 'C'->3, etc.)
        elif engine == "xlsxwriter":
            worksheet.set_column(col, col, max_col_len)

# get character length
def len_byte(value):
    length = len(value)
    utf8_length = len(value.encode('utf-8'))
    length = (utf8_length - length) / 1.2 + length # 根據excel實際欄寬調整參數
    return int(utf8_length)


if __name__ == "__main__":
    main()