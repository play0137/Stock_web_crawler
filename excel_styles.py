from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd
import pdb

file_path = "C:/Users/play0/OneDrive/桌面/stock/excel_styles.xlsx"
pf = PatternFill(fill_type="solid", start_color="FF0000")
# pdb.set_trace()

# load worksheets to get intersections
data_list = list()
dfs = pd.read_excel(file_path, sheet_name=None, header=None)
for sheet_name in dfs:
    df = dfs[sheet_name]
    data_list.append(set(df[0].values))
data_intersections = set.intersection(*data_list)

# fill the background color if the cell value in intersections
wb = load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row):
        for cell in row:
            if cell.value in data_intersections:
                cell.fill = pf

wb.save(file_path)