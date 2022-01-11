from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from openpyxl import Workbook

# styles setting
ft = Font(color="2D3EB9")
pf = PatternFill(fill_type="solid", start_color="2DB95E")

wb = Workbook()
ws = wb.active

a1 = ws['A1']
a1.value = 11
d4 = ws['D4']
d4.value = 4

a1.fill = pf
a1.font = ft
d4.font = ft

wb.save("C:/Users/play0/OneDrive/桌面/excel_styles.xlsx")